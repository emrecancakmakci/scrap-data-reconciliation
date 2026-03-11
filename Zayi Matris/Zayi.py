
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog, messagebox


def process_files():

    axa_path = filedialog.askopenfilename(title="Select Axata file", filetypes=[("Excel files", "*.xlsx")])
    sap_path = filedialog.askopenfilename(title="Select SAP file", filetypes=[("Excel files", "*.xlsx")])

    if not axa_path or not sap_path :
        messagebox.showerror("Error", "You must select the all files! ")
        return

    # Due to the Turkish interface output, the names of the 'df columns' are in Turkish in these lines.
    axa = pd.read_excel(axa_path)
    axa["Çıkış tarihi"] = pd.to_datetime(axa["Çıkış tarihi"], format="%d.%m.%Y %H:%M:%S", dayfirst=True)
    axa["Tarih"] = axa["Çıkış tarihi"].dt.date
    axa = axa[["Tarih", "SKU kodu", "SKU tanımı", "Miktar"]]


    axa_names = axa[["SKU kodu", "SKU tanımı"]].drop_duplicates().set_index("SKU kodu")

    axa_pivot = pd.pivot_table(
        axa,
        index="SKU kodu",
        values="Miktar",
        columns="Tarih",
        aggfunc="sum",
        margins=True,
        margins_name="Toplam Axa_a",
        fill_value=0
    )


    sap = pd.read_excel(sap_path)
    sap["Malzeme"] = sap["Malzeme"].str.replace("-", "")
    sap["Malzeme"] = sap["Malzeme"].str[1:]
    sap["Malzeme"] = sap["Malzeme"].astype("Int64")
    sap["Tarih"] = sap["Giriş tarihi"].dt.date
    sap = sap.rename(columns={"Malzeme": "SKU kodu", "Malzeme kısa metni": "SKU tanımı"})
    sap = sap[["Tarih", "SKU kodu", "SKU tanımı", "Miktar"]]


    sap_names = sap[["SKU kodu", "SKU tanımı"]].drop_duplicates().set_index("SKU kodu")

    sap_pivot = pd.pivot_table(
        sap,
        index="SKU kodu",
        values="Miktar",
        columns="Tarih",
        aggfunc="sum",
        margins=True,
        margins_name="Toplam SAP_s",
        fill_value=0
    )

    # Merged here
    merged = axa_pivot.join(sap_pivot, how="outer", lsuffix="_a", rsuffix="_s").fillna(0)

    merged = merged.join(axa_names, how="left")
    merged = merged.join(sap_names, how="left", rsuffix="_sap")

    merged["SKU tanımı"] = merged["SKU tanımı"].fillna(merged["SKU tanımı_sap"])
    merged = merged.drop(columns=["SKU tanımı_sap"])

    cols = merged.columns.tolist()
    cols = ["SKU tanımı"] + [c for c in cols if c != "SKU tanımı"]
    merged = merged[cols]

    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel files", "*.xlsx")],
                                             title="Select the file to save ")
    if not save_path:
        return

    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
        merged.to_excel(writer, sheet_name="Comparison")


    wb = load_workbook(save_path)
    ws = wb["Comparison"]

    blue_fill = PatternFill(start_color="79BBFC", end_color="79BBFC", fill_type="solid")
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for cell in ws[1]:
        if "_a" in str(cell.value):
            cell.fill = blue_fill
        elif "_s" in str(cell.value):
            cell.fill = grey_fill

    wb.save(save_path)
    messagebox.showinfo("Successful", " A comparison matrix was created and saved. ")
    root.destroy()


root = tk.Tk()
root.title("Create a loss/output comparison matrix.")
root.geometry("400x200")

label = tk.Label(root, text="Select the SAP and Axata loss/output log files. ",
                 wraplength=350)
label.pack(pady=20)

button = tk.Button(root, text="Select the files and create a comparison matrix", command=process_files)
button.pack(pady=20)

root.mainloop()